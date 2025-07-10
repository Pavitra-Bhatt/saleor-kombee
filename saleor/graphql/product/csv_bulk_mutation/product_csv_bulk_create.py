import openpyxl
from io import BytesIO
import graphene
from ...core.mutations import BaseMutation
from ..bulk_mutations.product_bulk_create import ProductBulkCreate, ProductBulkResult
from ...core.types.upload import Upload
from ...core.types import ProductBulkCreateError
from ...core.fields import JSONString
from ....permission.enums import ProductPermissions
from ....attribute.models import Attribute, AttributeValue
from django.utils.text import slugify
import re
from django.utils import timezone
from ....product.models import ProductChannelListing
import datetime
from graphene import InputObjectType
from decimal import Decimal
from ....product.models import ProductType
# import DateTime

def get_attribute_mapping(row=None):
    """Fetch all attributes and create mappings for both product and variant attributes."""
    attributes = Attribute.objects.all()
    print('attributes :', attributes)
    product_attrs = {}
    variant_attrs = {}

    # Convert row keys to lowercase for case-insensitive comparison
    row_keys = {k: v for k, v in row.items()} if row else {}

    # Check if this is a variant row
    is_variant_row = "Variant ID" in row_keys or "variant_id" in row_keys

    for attr in attributes:
        attr_id = graphene.Node.to_global_id("Attribute", attr.id)

        # Check if attribute exists in row (case-insensitive)
        attr_name_lower = attr.name.lower()
        matching_key = next((k for k in row_keys.keys() if k.lower() == attr_name_lower), None)

        if matching_key:
            if is_variant_row:
                variant_attrs[matching_key] = attr_id
            else:
                product_attrs[matching_key] = attr_id

    return product_attrs, variant_attrs

def process_attributes(data, attribute_mapping, is_variant=False):
    """Process attributes from the data and convert them to the required format."""
    attributes = []

    for key, value in data.items():
        if key in attribute_mapping:
            # Handle both single values and lists
            values = value if isinstance(value, list) else [value]
            # Filter out None/empty values
            values = [v for v in values if v]

            if values:
                attributes.append({
                    "id": attribute_mapping[key],
                    "values": values
                })

    return attributes

def to_snake_case(name):
    name = name.strip()
    name = re.sub(r"[\s\-]+", "_", name)  # Convert spaces/hyphens to underscores
    name = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)  # camelCase to camel_case
    name = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", name)  # handle lowerUpper
    name = re.sub(r"_+", "_", name)  # Collapse multiple underscores
    return name.lower()

def map_excel_row_keys(row, is_variant=False):
    """Map row data to the expected format."""
    mapped = {}

    # Handle product data
    if not is_variant:
        # Basic fields
        # if "Product ID" in row:
        #     mapped["product_reference"] = row["Product ID"]
        if "Product Name" in row:
            mapped["name"] = row["Product Name"]
        if "Product Type" in row:
            # Convert product type to global ID
            # product_data["product_type"] = graphene.Node.to_global_id("ProductType", product_type_obj.id)

            product_type = row["Product Type"]
            product_type_obj = ProductType.objects.get(name=product_type)
            print('product_type_obj :', product_type_obj)
            if hasattr(product_type, 'id'):
                mapped["product_type"] = graphene.Node.to_global_id("ProductType", product_type_obj.id)
            else:
                mapped["product_type"] = product_type

        # Handle channel listings
        if "Channel List" in row:
            channel_list = row["Channel List"]
            # Get the first channel if multiple are provided
            if isinstance(channel_list, str) and ',' in channel_list:
                channel_list = channel_list.split(',')[0].strip()

            # Get channel ID from database
            from ....channel.models import Channel
            try:
                channel = Channel.objects.first()  # Get the first channel
                if channel:
                    channel_id = graphene.Node.to_global_id("Channel", channel.id)
                    mapped["channel_listings"] = [{
                        "channel_id": channel_id,
                        "is_published": True,
                        "visible_in_listings": True,
                        "is_available_for_purchase": True,
                        "available_for_purchase_at": datetime.datetime.now(tz=datetime.UTC)
                    }]
            except Exception as e:
                print(f"Error getting channel: {e}")

        # Handle media
        media = []
        if "Thumbnail" in row and row["Thumbnail"]:
            media.append({
                "media_url": row["Thumbnail"],
                "alt": row.get("Product Name", "")
            })
        if "Thumbnail Hovered" in row and row["Thumbnail Hovered"]:
            media.append({
                "media_url": row["Thumbnail Hovered"],
                "alt": f"{row.get('Product Name', '')} - Hover"
            })
        if media:
            mapped["media"] = media

        # Process product attributes
        product_attrs, _ = get_attribute_mapping(row)
        if product_attrs:
            attributes = []
            for attr_name, attr_id in product_attrs.items():
                if attr_name in row and row[attr_name]:
                    # Handle multi-select: split comma-separated values into a list
                    raw_value = row[attr_name]
                    if isinstance(raw_value, str) and ',' in raw_value:
                        value_list = [v.strip() for v in raw_value.split(',') if v.strip()]
                    else:
                        value_list = [raw_value] if raw_value else []

                    # -------------------------------------------------------------
                    # MODIFIED: Create new attribute structure based on input style
                    # -------------------------------------------------------------
                    if attr_id and value_list:
                        # Get the actual attribute object to check its input_type
                        try:
                            # Extract the actual attribute ID from the global ID
                            attr_pk = graphene.Node.from_global_id(attr_id)[1]
                            attribute = Attribute.objects.get(pk=attr_pk)

                            # Process each value in the list
                            for value in value_list:
                                # Get the AttributeValue object for this value
                                try:
                                    attr_value = AttributeValue.objects.get(
                                        attribute=attribute,
                                        name=value
                                    )
                                    attr_value_id = graphene.Node.to_global_id("AttributeValue", attr_value.id)

                                    # Create attribute structure based on input_type
                                    attribute_data = {
                                        "id": attr_id
                                    }

                                    # Map input_type to the corresponding field name
                                    input_type_mapping = {
                                        'dropdown': 'dropdown',
                                        'multiselect': 'multiselect',
                                        'swatch': 'swatch',
                                        'boolean': 'boolean',
                                        'date': 'date',
                                        'date_time': 'date_time',
                                        'numeric': 'numeric',
                                        'rich_text': 'rich_text',
                                        'plain_text': 'plain_text',
                                        'reference': 'reference',
                                        'file': 'file'
                                    }

                                    field_name = input_type_mapping.get(attribute.input_type, 'dropdown')
                                    attribute_data[field_name] = {
                                        "id": attr_value_id
                                    }

                                    attributes.append(attribute_data)

                                except AttributeValue.DoesNotExist:
                                    print(f"AttributeValue '{value}' not found for attribute '{attr_name}'")
                                    continue

                        except (Attribute.DoesNotExist, Exception) as e:
                            print(f"Error processing attribute '{attr_name}': {e}")
                            continue
                    # -------------------------------------------------------------
                    # END MODIFICATION
                    # -------------------------------------------------------------
            if attributes:
                mapped["attributes"] = attributes

    # Handle variant data
    else:
        # if "Variant ID" in row:
        #     mapped["variant_reference"] = row["Variant ID"]
        # if "Product ID" in row:
        #     mapped["product_reference"] = row["Product ID"]
        if "Variant Name" in row:
            mapped["name"] = row["Variant Name"]
        if "SKU" in row:
            mapped["sku"] = row["SKU"]

        # Handle variant channel listings
        if "Price" in row or "Cost Price" in row:
            # Get channel ID from database
            from ....channel.models import Channel
            try:
                channel = Channel.objects.first()  # Get the first channel
                if channel:
                    channel_id = graphene.Node.to_global_id("Channel", channel.id)
                    mapped["channel_listings"] = [{
                        "channel_id": channel_id,
                        "price": Decimal(str(row.get("Price", 0))),
                        "cost_price": Decimal(str(row.get("Cost Price", 0)))
                    }]
            except Exception as e:
                print(f"Error getting channel: {e}")

        # Handle stocks
        if "Inventory" in row:
            # Get warehouse ID from database
            from ....warehouse.models import Warehouse
            try:
                warehouse = Warehouse.objects.first()  # Get the first warehouse
                if warehouse:
                    warehouse_id = graphene.Node.to_global_id("Warehouse", warehouse.id)
                    mapped["stocks"] = [{
                        "warehouse": warehouse_id,
                        "quantity": int(row["Inventory"])
                    }]
            except Exception as e:
                print(f"Error getting warehouse: {e}")

        # Process variant attributes
        _, variant_attrs = get_attribute_mapping(row)
        if variant_attrs:
            attributes = []
            for attr_name, attr_id in variant_attrs.items():
                if attr_name in row and row[attr_name]:
                    # -------------------------------------------------------------
                    # MODIFIED: Create new attribute structure for variants
                    # -------------------------------------------------------------
                    try:
                        # Extract the actual attribute ID from the global ID
                        attr_pk = graphene.Node.from_global_id(attr_id)[1]
                        attribute = Attribute.objects.get(pk=attr_pk)

                        # Get the AttributeValue object for this value
                        try:
                            attr_value = AttributeValue.objects.get(
                                attribute=attribute,
                                name=str(row[attr_name])
                            )
                            attr_value_id = graphene.Node.to_global_id("AttributeValue", attr_value.id)

                            # Create attribute structure based on input_type
                            attribute_data = {
                                "id": attr_id
                            }

                            # Map input_type to the corresponding field name
                            input_type_mapping = {
                                'dropdown': 'dropdown',
                                'multiselect': 'multiselect',
                                'swatch': 'swatch',
                                'boolean': 'boolean',
                                'date': 'date',
                                'date_time': 'date_time',
                                'numeric': 'numeric',
                                'rich_text': 'rich_text',
                                'plain_text': 'plain_text',
                                'reference': 'reference',
                                'file': 'file'
                            }

                            field_name = input_type_mapping.get(attribute.input_type, 'dropdown')
                            attribute_data[field_name] = {
                                "id": attr_value_id
                            }

                            attributes.append(attribute_data)

                        except AttributeValue.DoesNotExist:
                            print(f"AttributeValue '{row[attr_name]}' not found for attribute '{attr_name}'")
                            continue

                    except (Attribute.DoesNotExist, Exception) as e:
                        print(f"Error processing variant attribute '{attr_name}': {e}")
                        continue
                    # -------------------------------------------------------------
                    # END MODIFICATION
                    # -------------------------------------------------------------
            if attributes:
                mapped["attributes"] = attributes

    return mapped

def get_or_create_attribute_and_value(attr_name, value, is_variant=False):
    """Get attribute and its value, only querying existing ones."""
    # 1. Get the attribute
    attribute = Attribute.objects.filter(name=attr_name).first()
    if not attribute:
        print(f"Attribute '{attr_name}' not found")
        return None, None

    # 2. Get the attribute value
    attr_value = AttributeValue.objects.filter(
        attribute=attribute,
        name=value
    ).first()

    if not attr_value:
        print(f"Attribute value '{value}' not found for attribute '{attr_name}'")
        return None, None

    # 3. Return global IDs for GraphQL
    from graphene import Node
    attr_id = Node.to_global_id("Attribute", attribute.id)
    value_id = Node.to_global_id("AttributeValue", attr_value.id)

    return attr_id, value_id

class ProductChannelListingCreateInput(InputObjectType):
    channel_id = graphene.ID(required=True)
    is_published = graphene.Boolean()
    published_at = graphene.DateTime()
    visible_in_listings = graphene.Boolean()
    is_available_for_purchase = graphene.Boolean()
    available_for_purchase_at = graphene.DateTime()

class ProductBulkCreateFromExcel(BaseMutation):
    count = graphene.Int()
    results = graphene.List(ProductBulkResult)

    class Arguments:
        excel_file = Upload(required=False)
        products = graphene.List(JSONString, required=False)

    class Meta:
        description = "Bulk create products and variants from Excel or direct input"
        permissions = (ProductPermissions.MANAGE_PRODUCTS,)
        error_type_class = ProductBulkCreateError

    @classmethod
    def perform_mutation(cls, root, info, **data):
        user = getattr(getattr(info, "context", None), "user", None)
        if user is None:
            raise Exception("User not found in request context")
        all_attributes = Attribute.objects.all()
        for attr in all_attributes:
            print(f"ID: {attr.id}, Name: {attr.name}, Slug: {attr.slug}, Type: {attr.type}, Input Type: {attr.input_type}")

        products_input = None

        if data.get("products"):
            # Handle direct JSON input
            products_data = data["products"]
            products_input = []

            for product_data in products_data:
                # Map the product data to the expected format
                mapped_product = map_excel_row_keys(product_data, is_variant=False)
                products_input.append(mapped_product)

        elif data.get("excel_file"):
            excel_file = data["excel_file"]
            try:
                if hasattr(excel_file, "read"):
                    file_content = excel_file.read()
                elif isinstance(excel_file, bytes):
                    file_content = excel_file
                elif isinstance(excel_file, str):
                    try:
                        with open(excel_file, "rb") as f:
                            file_content = f.read()
                        # If it's a file path, ensure to rewind for openpyxl
                        # f.seek(0)
                    except Exception:
                        import base64
                        try:
                            file_content = base64.b64decode(excel_file)
                        except Exception:
                            raise Exception("Excel file is not a valid file, path, or base64 string")
                else:
                    raise Exception("Unsupported file type for Excel upload")

                wb = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
            except Exception as e:
                raise Exception(f"Failed to load Excel file: {str(e)}")

            products = []
            variants = []

            if "Products" in wb.sheetnames:
                sheet = wb["Products"]
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    raw_row = dict(zip(headers, row))
                    products.append(map_excel_row_keys(raw_row, is_variant=False))
            else:
                raise Exception("Sheet 'products' not found in Excel file.")

            if "Variants" in wb.sheetnames:
                sheet = wb["Variants"]
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    raw_row = dict(zip(headers, row))
                    variants.append(map_excel_row_keys(raw_row, is_variant=True))
            else:
                raise Exception("Sheet 'variants' not found in Excel file.")
            print('products :', products)
            print('variants :', variants)
            # Map variants to their products by product_reference
            # product_map = {p["product_reference"]: p for p in products}
            # for variant in variants:
            #     ref = variant.get("product_reference")
            #     if ref in product_map:
            #         product_map[ref].setdefault("variants", []).append(variant)
            # products_input = list(product_map.values())
            # print('products_input 12:', products_input)
        else:
            raise Exception("No valid input provided. Provide either 'excel_file' or 'products'.")

        # Process channel listings for each product
        for product_data in products_input:
            if "channel_listings" in product_data:
                listings_to_create = []
                updated_channels = set()
                cls.prepare_products_channel_listings(
                    product_data,
                    product_data["channel_listings"],
                    listings_to_create,
                    updated_channels,
                )
        print('products_input :', products_input)
        return ProductBulkCreate.perform_mutation(root, info, products=products_input)

    @classmethod
    def prepare_products_channel_listings(cls, product_data, listings_input, listings_to_create, updated_channels):
        """Only query existing channel listings without creating new ones."""
        from ....channel.models import Channel
        from ....product.models import Product, ProductChannelListing
        from graphene import Node

        # Query existing product
        product = None
        # if "product_reference" in product_data:
        #     try:
        #         product = Product.objects.get(product_reference=product_data["product_reference"])
        #     except Product.DoesNotExist:
        #         print(f"Product with reference {product_data['product_reference']} not found")
        #         return

        # Query existing channel listings
        if product:
            existing_listings = ProductChannelListing.objects.filter(product=product)
            for listing in existing_listings:
                listings_to_create.append(listing)
                updated_channels.add(listing.channel)
